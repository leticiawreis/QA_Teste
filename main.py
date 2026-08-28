import sys
import os
import shutil
from pathlib import Path
from PyQt6.QtCore import QDate
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLineEdit, QTextEdit, QComboBox, QDateEdit, QPushButton, QLabel,
    QStackedWidget, QListWidget, QListWidgetItem, QMessageBox, QFileDialog,
    QTableWidget, QTableWidgetItem, QFrame
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------

# Pasta onde fica tudo: as planilhas de cada categoria e as evidências anexadas.
# Usa a pasta onde o .exe (ou este .py, se rodado direto) está localizado —
# assim o executável e os dados sempre ficam juntos, não importa pra onde
# você mova a pasta inteira.
if getattr(sys, "frozen", False):
    # Rodando como .exe gerado pelo PyInstaller
    BASE_DIR = Path(sys.executable).parent
else:
    # Rodando como script .py normal
    BASE_DIR = Path(__file__).resolve().parent

APP_DIR = BASE_DIR
QA_DIR = APP_DIR / "QA"

FOLDERS = [
    "01 - Dashboard", "02 - Casos de Teste", "03 - Bugs", "04 - Evidências",
    "05 - Base de Conhecimento", "06 - Relatórios Mensais", "07 - Riscos",
    "08 - Checklists",
]

# Cada categoria tem UM único Excel (não um arquivo por robô/RPA).
# Toda vez que algo é registrado na interface, uma nova linha é adicionada
# automaticamente nesse Excel.
EXCEL_SHEETS = {
    "tests": {
        "folder": "02 - Casos de Teste",
        "file": "Casos de Teste.xlsx",
        "headers": ["ID", "RPA", "Responsável", "Data", "O que o RPA faz",
                    "Cenário / passos", "Resultado", "Erros / observações", "Evidência"],
    },
    "bugs": {
        "folder": "03 - Bugs",
        "file": "Bugs.xlsx",
        "headers": ["ID", "RPA", "Responsável", "Data", "Descrição",
                    "Severidade", "Status", "Evidência"],
    },
    "risks": {
        "folder": "07 - Riscos",
        "file": "Riscos.xlsx",
        "headers": ["ID", "RPA / Título", "Responsável", "Data", "Descrição",
                    "Impacto", "Probabilidade", "Status"],
    },
    "knowledge": {
        "folder": "05 - Base de Conhecimento",
        "file": "Base de Conhecimento.xlsx",
        "headers": ["ID", "Título", "Categoria", "Descrição", "Solução", "Data"],
    },
    "checklists": {
        "folder": "08 - Checklists",
        "file": "Checklists.xlsx",
        "headers": ["ID", "RPA / Título", "Responsável", "Data", "Itens", "Resultado"],
    },
}


def excel_path(table: str) -> Path:
    spec = EXCEL_SHEETS[table]
    return QA_DIR / spec["folder"] / spec["file"]


def init_storage():
    """Cria as pastas do projeto e, para cada categoria, o Excel com cabeçalho
    (se ele ainda não existir)."""
    APP_DIR.mkdir(parents=True, exist_ok=True)
    QA_DIR.mkdir(exist_ok=True)
    for folder in FOLDERS:
        (QA_DIR / folder).mkdir(exist_ok=True)

    for table, spec in EXCEL_SHEETS.items():
        path = excel_path(table)
        if not path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = spec["file"].replace(".xlsx", "")[:31]
            ws.append(spec["headers"])
            header_fill = PatternFill("solid", fgColor="1B1B1F")
            header_font = Font(color="FF9F40", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col_index, header in enumerate(spec["headers"], start=1):
                ws.column_dimensions[ws.cell(row=1, column=col_index).column_letter].width = max(14, len(header) + 2)
            wb.save(path)


def append_record(table: str, values: list) -> int:
    """Adiciona uma linha ao Excel da categoria e devolve o ID gerado (sequencial)."""
    path = excel_path(table)
    wb = load_workbook(path)
    ws = wb.active
    ident = ws.max_row  # linha 1 é cabeçalho, então isso já é o próximo ID (1-based)
    ws.append([ident, *values])
    wb.save(path)
    return ident


def read_records(table: str) -> list:
    """Lê todas as linhas (sem cabeçalho) da planilha da categoria."""
    path = excel_path(table)
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[0] is not None]
    wb.close()
    return rows


def copy_evidence(evidence_path: str, record_id: str):
    """Copia o arquivo de evidência para a pasta 04 - Evidências/<record_id>/."""
    if not evidence_path:
        return
    dest = QA_DIR / "04 - Evidências" / record_id
    dest.mkdir(parents=True, exist_ok=True)
    shutil.copy2(evidence_path, dest / Path(evidence_path).name)


# ---------------------------------------------------------------------------
# Widgets auxiliares
# ---------------------------------------------------------------------------

class Card(QFrame):
    """Cartão de indicador exibido no dashboard."""

    def __init__(self, title, value):
        super().__init__()
        self.setObjectName("card")
        layout = QVBoxLayout(self)
        title_label = QLabel(title)
        title_label.setObjectName("cardTitle")
        value_label = QLabel(str(value))
        value_label.setObjectName("cardValue")
        layout.addWidget(title_label)
        layout.addWidget(value_label)


class FormBase(QWidget):
    """Funcionalidades comuns a todos os formulários."""

    def msg(self, text):
        QMessageBox.information(self, "QA Control Center", text)

    def error(self, text):
        QMessageBox.critical(self, "Erro", text)

    def build_title(self, layout, text):
        label = QLabel(text)
        label.setObjectName("title")
        layout.addWidget(label)


class EvidenceAttachment(QWidget):
    """Botão + label reutilizáveis para anexar um arquivo de evidência."""

    def __init__(self):
        super().__init__()
        self.path = ""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        self.label = QLabel("Nenhum arquivo anexado.")
        button = QPushButton("📎 Anexar evidência")
        button.clicked.connect(self._pick_file)
        layout.addWidget(button)
        layout.addWidget(self.label)

    def _pick_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Selecionar evidência")
        if path:
            self.path = path
            self.label.setText(Path(path).name)


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

class Dashboard(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        title = QLabel("📊 Dashboard de QA")
        title.setObjectName("title")
        layout.addWidget(title)

        self.cards_row = QHBoxLayout()
        layout.addLayout(self.cards_row)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["ID", "RPA", "Data", "Resultado", "Erros"])
        layout.addWidget(self.table)

        self.refresh()

    def refresh(self):
        while self.cards_row.count():
            item = self.cards_row.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        tests = read_records("tests")
        bugs = read_records("bugs")

        total = len(tests)
        approved = sum(1 for r in tests if r[6] == "Aprovado")   # coluna Resultado
        failed = sum(1 for r in tests if r[6] == "Reprovado")
        open_bugs = sum(1 for r in bugs if r[6] != "Resolvido")  # coluna Status

        for title, value in [
            ("Testes", total), ("Aprovados", approved),
            ("Reprovados", failed), ("Bugs abertos", open_bugs),
        ]:
            self.cards_row.addWidget(Card(title, value))

        last_tests = list(reversed(tests))[:20]
        self.table.setRowCount(len(last_tests))
        for row_index, row in enumerate(last_tests):
            # row: ID, RPA, Responsável, Data, Descrição, Cenário, Resultado, Erros, Evidência
            display = [row[0], row[1], row[3], row[6], row[7]]
            for col_index, value in enumerate(display):
                self.table.setItem(row_index, col_index, QTableWidgetItem(str(value or "")))


# ---------------------------------------------------------------------------
# Formulário de Casos de Teste
# ---------------------------------------------------------------------------

class TestForm(FormBase):
    def __init__(self, refresh_dashboard):
        super().__init__()
        self.refresh_dashboard = refresh_dashboard

        layout = QVBoxLayout(self)
        self.build_title(layout, "🧪 Registrar Caso de Teste")

        form = QFormLayout()
        self.rpa = QLineEdit()
        self.responsible = QLineEdit()
        self.date = QDateEdit(QDate.currentDate())
        self.date.setCalendarPopup(True)
        self.description = QTextEdit()
        self.scenario = QTextEdit()
        self.result = QComboBox()
        self.result.addItems(["Aprovado", "Reprovado", "Bloqueado", "Não executado"])
        self.errors = QTextEdit()

        form.addRow("RPA:", self.rpa)
        form.addRow("Responsável:", self.responsible)
        form.addRow("Data:", self.date)
        form.addRow("O que o RPA faz:", self.description)
        form.addRow("Cenário / passos:", self.scenario)
        form.addRow("Resultado:", self.result)
        form.addRow("Erros / observações:", self.errors)
        layout.addLayout(form)

        self.evidence = EvidenceAttachment()
        layout.addWidget(self.evidence)

        save_button = QPushButton("✓ REGISTRAR TESTE")
        save_button.clicked.connect(self.save)
        layout.addWidget(save_button)

    def save(self):
        if not self.rpa.text().strip() or not self.responsible.text().strip():
            return self.error("Preencha pelo menos RPA e Responsável.")

        date_str = self.date.date().toString("dd/MM/yyyy")
        ident = append_record("tests", [
            self.rpa.text(), self.responsible.text(), date_str,
            self.description.toPlainText(), self.scenario.toPlainText(),
            self.result.currentText(), self.errors.toPlainText(), self.evidence.path,
        ])
        copy_evidence(self.evidence.path, f"CT-{ident:05d}")

        self.msg(f"Teste CT-{ident:05d} registrado com sucesso.")
        self.refresh_dashboard()


# ---------------------------------------------------------------------------
# Formulário de Bugs
# ---------------------------------------------------------------------------

class BugForm(FormBase):
    def __init__(self, refresh_dashboard):
        super().__init__()
        self.refresh_dashboard = refresh_dashboard

        layout = QVBoxLayout(self)
        self.build_title(layout, "🐞 Registrar Bug")

        form = QFormLayout()
        self.rpa = QLineEdit()
        self.responsible = QLineEdit()
        self.date = QDateEdit(QDate.currentDate())
        self.date.setCalendarPopup(True)
        self.description = QTextEdit()
        self.severity = QComboBox()
        self.severity.addItems(["Crítica", "Alta", "Média", "Baixa"])
        self.status = QComboBox()
        self.status.addItems(["Aberto", "Em análise", "Corrigido", "Resolvido", "Reaberto"])

        form.addRow("RPA:", self.rpa)
        form.addRow("Responsável:", self.responsible)
        form.addRow("Data:", self.date)
        form.addRow("Descrição:", self.description)
        form.addRow("Severidade:", self.severity)
        form.addRow("Status:", self.status)
        layout.addLayout(form)

        self.evidence = EvidenceAttachment()
        layout.addWidget(self.evidence)

        save_button = QPushButton("✓ REGISTRAR BUG")
        save_button.clicked.connect(self.save)
        layout.addWidget(save_button)

    def save(self):
        if not self.rpa.text().strip() or not self.description.toPlainText().strip():
            return self.error("Preencha RPA e descrição.")

        date_str = self.date.date().toString("dd/MM/yyyy")
        ident = append_record("bugs", [
            self.rpa.text(), self.responsible.text(), date_str,
            self.description.toPlainText(), self.severity.currentText(),
            self.status.currentText(), self.evidence.path,
        ])
        copy_evidence(self.evidence.path, f"BUG-{ident:05d}")

        self.msg(f"BUG-{ident:05d} registrado.")
        self.refresh_dashboard()


# ---------------------------------------------------------------------------
# Formulário genérico (Riscos, Base de Conhecimento, Checklists)
# ---------------------------------------------------------------------------

class SimpleForm(FormBase):
    """
    Formulário reutilizável para categorias simples (riscos, conhecimento, checklists).
    `spec` define título, tabela (chave em EXCEL_SHEETS), prefixo do ID e combos extras.
    """

    def __init__(self, spec, refresh_dashboard):
        super().__init__()
        self.spec = spec
        self.refresh_dashboard = refresh_dashboard

        layout = QVBoxLayout(self)
        self.build_title(layout, spec["title"])

        form = QFormLayout()
        self.rpa = QLineEdit()
        self.responsible = QLineEdit()
        self.date = QDateEdit(QDate.currentDate())
        self.date.setCalendarPopup(True)
        self.description = QTextEdit()

        form.addRow("RPA / título:", self.rpa)
        form.addRow("Responsável:", self.responsible)
        form.addRow("Data:", self.date)
        form.addRow("Descrição:", self.description)

        self.combos = []
        for label, items in spec.get("combos", []):
            combo = QComboBox()
            combo.addItems(items)
            self.combos.append(combo)
            form.addRow(label, combo)

        layout.addLayout(form)

        save_button = QPushButton("✓ REGISTRAR")
        save_button.clicked.connect(self.save)
        layout.addWidget(save_button)

    def save(self):
        table = self.spec["table"]
        date_str = self.date.date().toString("dd/MM/yyyy")
        description = self.description.toPlainText()

        if table == "risks":
            impact, probability, status = (c.currentText() for c in self.combos)
            values = [self.rpa.text(), self.responsible.text(), date_str, description,
                      impact, probability, status]
        elif table == "knowledge":
            category = self.combos[0].currentText()
            values = [self.rpa.text(), category, description, "", date_str]
        else:  # checklists
            result = self.combos[0].currentText()
            values = [self.rpa.text(), self.responsible.text(), date_str, description, result]

        ident = append_record(table, values)

        self.msg(f"Registro {self.spec['prefix']}-{ident:05d} criado.")
        self.refresh_dashboard()


# ---------------------------------------------------------------------------
# Janela principal
# ---------------------------------------------------------------------------

NAV_ITEMS = [
    "📊 Dashboard", "🧪 Casos de Teste", "🐞 Bugs", "⚠️ Riscos",
    "📚 Conhecimento", "📋 Checklists", "📁 Abrir pasta QA",
]

FORM_SPECS = [
    {
        "title": "⚠️ Registrar Risco", "table": "risks", "prefix": "RISK",
        "combos": [
            ("Impacto:", ["Alto", "Médio", "Baixo"]),
            ("Probabilidade:", ["Alta", "Média", "Baixa"]),
            ("Status:", ["Aberto", "Monitorando", "Mitigado", "Encerrado"]),
        ],
    },
    {
        "title": "📚 Base de Conhecimento", "table": "knowledge", "prefix": "KB",
        "combos": [("Categoria:", ["Bug recorrente", "Solução", "Processo", "Dica", "Outro"])],
    },
    {
        "title": "📋 Checklist", "table": "checklists", "prefix": "CHK",
        "combos": [("Resultado:", ["Aprovado", "Reprovado", "Com pendências"])],
    },
]


class Main(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("QA Control Center")
        self.resize(1100, 700)

        root = QWidget()
        main_layout = QHBoxLayout(root)
        self.setCentralWidget(root)

        self.nav = QListWidget()
        self.nav.setFixedWidth(220)
        for item in NAV_ITEMS:
            self.nav.addItem(QListWidgetItem(item))
        main_layout.addWidget(self.nav)

        self.stack = QStackedWidget()
        main_layout.addWidget(self.stack)

        self.dashboard = Dashboard()
        self.stack.addWidget(self.dashboard)
        self.stack.addWidget(TestForm(self.refresh))
        self.stack.addWidget(BugForm(self.refresh))
        for spec in FORM_SPECS:
            self.stack.addWidget(SimpleForm(spec, self.refresh))

        self.nav.currentRowChanged.connect(self.change_page)
        self.nav.setCurrentRow(0)

    def change_page(self, row):
        if row < self.stack.count():
            self.stack.setCurrentIndex(row)
        else:
            os.startfile(str(QA_DIR.resolve()))

    def refresh(self):
        self.dashboard.refresh()


# ---------------------------------------------------------------------------
# Tema escuro moderno com detalhes em âmbar
# ---------------------------------------------------------------------------

STYLE_SHEET = """
QWidget { font-family: 'Segoe UI'; font-size: 13.5px; color: #F2F2F3; }
QMainWindow { background: #121214; }

QListWidget {
    background: #0B0B0D;
    border: 0;
    border-right: 1px solid #232327;
    padding: 16px 10px;
    outline: 0;
}
QListWidget::item {
    padding: 12px 14px;
    border-radius: 10px;
    margin-bottom: 2px;
    color: #B9B9C0;
}
QListWidget::item:selected {
    background: rgba(255, 159, 64, 0.14);
    color: #FF9F40;
    font-weight: 600;
}
QListWidget::item:hover:!selected {
    background: #18181C;
    color: #F2F2F3;
}

QStackedWidget { background: #121214; }

QLineEdit, QTextEdit, QComboBox, QDateEdit {
    padding: 9px 11px;
    border: 1px solid #2B2B31;
    border-radius: 10px;
    background: #1B1B1F;
    color: #F2F2F3;
    selection-background-color: #FF9F40;
    selection-color: #121214;
}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus, QDateEdit:focus {
    border: 1px solid #FF9F40;
}
QComboBox::drop-down { border: 0; width: 24px; }
QComboBox QAbstractItemView {
    background: #1B1B1F;
    color: #F2F2F3;
    border: 1px solid #2B2B31;
    selection-background-color: rgba(255, 159, 64, 0.18);
    selection-color: #FF9F40;
    outline: 0;
}
QLabel { color: #D4D4D8; }

QPushButton {
    padding: 11px 18px;
    border-radius: 10px;
    background: #FF9F40;
    color: #121214;
    font-weight: 600;
    border: 0;
}
QPushButton:hover { background: #FFB566; }
QPushButton:pressed { background: #E8862A; }

QLabel#title {
    font-size: 22px;
    font-weight: 700;
    color: #F2F2F3;
    margin-bottom: 10px;
}

QFrame#card {
    background: #1B1B1F;
    border: 1px solid #2B2B31;
    border-radius: 14px;
    min-width: 170px;
    padding: 4px;
}
QLabel#cardTitle { color: #9A9AA4; font-size: 12.5px; }
QLabel#cardValue { font-size: 28px; font-weight: 700; color: #FF9F40; }

QTableWidget {
    background: #1B1B1F;
    color: #F2F2F3;
    gridline-color: #232327;
    border: 1px solid #2B2B31;
    border-radius: 10px;
    selection-background-color: rgba(255, 159, 64, 0.18);
    selection-color: #F2F2F3;
}
QHeaderView::section {
    background: #0B0B0D;
    color: #9A9AA4;
    padding: 8px;
    border: 0;
    border-bottom: 1px solid #2B2B31;
    font-weight: 600;
}
QTableWidget::item { padding: 4px; }
QTableWidget::item:selected { background: rgba(255, 159, 64, 0.18); color: #F2F2F3; }

QScrollBar:vertical {
    background: transparent;
    width: 10px;
    margin: 0;
}
QScrollBar::handle:vertical {
    background: #2B2B31;
    border-radius: 5px;
    min-height: 24px;
}
QScrollBar::handle:vertical:hover { background: #3A3A42; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
"""


def main():
    init_storage()
    app = QApplication(sys.argv)
    app.setStyleSheet(STYLE_SHEET)
    window = Main()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()